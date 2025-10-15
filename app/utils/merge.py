from __future__ import annotations

import csv
import re
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from dateutil import parser as dateparser
from openpyxl import load_workbook


NORMALIZED_COLUMN_ALIASES: Dict[str, List[str]] = {
    # Key columns
    "organisation_id": [
        "organisation id",
        "org id",
        "org_id",
        "organisationid",
        "organization id",
        "organization_id",
        "company id",
        "company_id",
    ],
    "customer": ["customer", "client", "account", "buyer"],
    "name": ["name", "contact name", "full name", "contractor", "contractor name"],
    # Display fields
    "product_status": ["product status", "product_status", "status product"],
    "engagement_status": [
        "engagement status",
        "engagement_status",
        "status engagement",
        "project status",
    ],
    "carbon_factor": ["carbon factor", "carbon_factor", "co2 factor", "emissions factor"],
    "next_activity_due_date": [
        "next activity due date",
        "next due",
        "due date",
        "next_activity",
        "next action due",
    ],
}


def _canonicalize(name: str) -> str:
    s = name.strip().lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _map_column(name: str) -> str:
    c_norm = _canonicalize(str(name))
    for canonical, aliases in NORMALIZED_COLUMN_ALIASES.items():
        if c_norm == canonical or c_norm in aliases:
            return canonical
    return c_norm


def read_table(path: str) -> List[Dict[str, Any]]:
    if path.lower().endswith((".xlsx", ".xls")):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            return []
        headers = [str(h) if h is not None else "" for h in headers]
        records: List[Dict[str, Any]] = []
        for row in rows_iter:
            rec = {}
            for h, v in zip(headers, row):
                if h == "":
                    continue
                rec[h] = v
            # skip entirely empty rows
            if any(v is not None and str(v).strip() != "" for v in rec.values()):
                records.append(rec)
        return records
    else:
        # CSV
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            return [dict(row) for row in reader]


def normalize_records(records: List[Dict[str, Any]]) -> tuple[List[Dict[str, Any]], Dict[str, str]]:
    # Build column map from first record keys (and extend as new keys appear)
    col_map: Dict[str, str] = {}
    normalized: List[Dict[str, Any]] = []
    for rec in records:
        out: Dict[str, Any] = {}
        for k, v in rec.items():
            if k not in col_map:
                col_map[k] = _map_column(k)
            out[col_map[k]] = v
        normalized.append(out)
    return normalized, col_map


def infer_key_columns(frames: Iterable[List[Dict[str, Any]]], key_options: List[List[str]]) -> Optional[List[str]]:
    frames_list = list(frames)
    for keys in key_options:
        ok = True
        for frame in frames_list:
            cols = set()
            for rec in frame:
                cols.update(rec.keys())
            if not all(k in cols for k in keys):
                ok = False
                break
        if ok:
            return keys
    return None


def _compose_key(rec: Dict[str, Any], keys: Sequence[str]) -> Optional[Tuple[str, ...]]:
    values: List[str] = []
    for k in keys:
        v = rec.get(k)
        if v is None:
            return None
        s = str(v).strip()
        if s == "":
            return None
        values.append(s)
    return tuple(values)


def combine_records(frames: Iterable[List[Dict[str, Any]]], keys: List[str]) -> List[Dict[str, Any]]:
    frames_list = list(frames)
    combined_map: Dict[Tuple[str, ...], Dict[str, Any]] = {}
    for frame in frames_list:
        for rec in frame:
            k = _compose_key(rec, keys)
            if k is None:
                # Skip rows that lack join keys
                continue
            target = combined_map.setdefault(k, {})
            # Ensure key fields are present
            for i, key_name in enumerate(keys):
                target.setdefault(key_name, k[i])
            # Merge preferring existing non-empty, else new value
            for col, val in rec.items():
                if val is None or str(val).strip() == "":
                    continue
                if col not in target or target[col] is None or str(target[col]).strip() == "":
                    target[col] = val
    return list(combined_map.values())


def parse_date(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        # likely Excel serial is not handled here; leave as None
        try:
            return dateparser.parse(str(value)).date().isoformat()
        except Exception:
            return None
    s = str(value).strip()
    if not s:
        return None
    try:
        dt = dateparser.parse(s, dayfirst=False)
        return dt.date().isoformat()
    except Exception:
        return None
